"""
Excel Generator for Municipality Data
Creates Excel files with municipality data and equipment calculations
Supports multiple countries with localized labels
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os

# Handle path for both script and PyInstaller .exe
if getattr(sys, 'frozen', False):
    sys.path.insert(0, sys._MEIPASS)
else:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import (
    OUTPUT_FULL_EXCEL,
    OUTPUT_SIMPLE_EXCEL,
    COLUMN_NAMES,
    SIMPLE_COLUMNS,
    DATA_DIR,
    POPULATION_THRESHOLD,
    MIN_RURAL_POPULATION,
    EQUIPMENT_DIVISOR_URBAN,
    EQUIPMENT_DIVISOR_RURAL,
    DEFAULT_URBAN_PERCENTAGE
)


class ExcelGenerator:
    """Generator for Excel files with municipality data"""

    def __init__(self, country_config=None):
        """
        Initialize generator with optional country configuration

        Args:
            country_config: Dictionary with country-specific settings
        """
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set country-specific labels or use defaults
        if country_config and "labels" in country_config:
            labels = country_config["labels"]
            self.column_names = {
                'name': labels.get('municipality', COLUMN_NAMES['name']),
                'total_hab': labels.get('total_hab', COLUMN_NAMES['total_hab']),
                'hab_urban': labels.get('hab_urban', COLUMN_NAMES['hab_urban']),
                'hab_rural': labels.get('hab_rural', COLUMN_NAMES['hab_rural']),
                'equipos_urban': labels.get('equipos_urban', COLUMN_NAMES['equipos_urban']),
                'equipos_rural': labels.get('equipos_rural', COLUMN_NAMES['equipos_rural']),
                'total_equipos': labels.get('total_equipos', COLUMN_NAMES['total_equipos']),
            }
            self.simple_columns = {
                'name': labels.get('municipality', SIMPLE_COLUMNS['name']),
                'total_equipos': labels.get('total_equipos', SIMPLE_COLUMNS['total_equipos']),
            }
            self.country_name = country_config.get("name", "España")
        else:
            self.column_names = COLUMN_NAMES
            self.simple_columns = SIMPLE_COLUMNS
            self.country_name = "España"

        # Generate country-specific file names
        country_suffix = self._get_country_suffix()
        self.output_full = os.path.join(DATA_DIR, f"municipios_{country_suffix}_completo.xlsx")
        self.output_simple = os.path.join(DATA_DIR, f"municipios_{country_suffix}_clasificacion.xlsx")

    def _get_country_suffix(self):
        """Get lowercase country name for file naming"""
        # Map country names to simple file suffixes
        country_map = {
            "España": "espana",
            "France": "france",
            "Italia": "italia",
            "Portugal": "portugal",
            "Deutschland": "deutschland",
        }
        return country_map.get(self.country_name, self.country_name.lower().replace(" ", "_"))

    def _ensure_data_dir(self):
        """Ensure data directory exists"""
        os.makedirs(DATA_DIR, exist_ok=True)

    def _apply_header_style(self, ws, row_num=1):
        """Apply styling to header row"""
        for cell in ws[row_num]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _calculate_equipment_data(self, municipalities):
        """
        Calculate equipment data for each municipality

        Logic:
        - If population <= 2000 (Rural):
            - HAB RURAL = 100% of population
            - HAB URBAN = empty
            - EQUIPOS RURAL = population / 51
            - EQUIPOS URBAN = 0

        - If population > 2000 (Urban):
            - Fixed base: 2000 (fijo) -> base_equipos = 2000 / 51
            - Excess population (population - 2000) is distributed by IA ratio:
                - HAB URBAN = excess * urban_percentage
                - HAB RURAL (extra) = excess * (1 - urban_percentage)
            - EQUIPOS URBAN = hab_urban / 301
            - EQUIPOS RURAL (extra) = hab_rural_extra / 51
            - TOTAL EQUIPOS = base_equipos + equipos_urban + equipos_rural_extra
        """
        processed = []

        for m in municipalities:
            population = m.get('population') or 0
            name = m.get('name', '')
            # Get IA classification ratio (urban_percentage), default to config value
            urban_percentage = m.get('urban_percentage', DEFAULT_URBAN_PERCENTAGE)

            # Determine if exceeds threshold
            exceeds_threshold = population > POPULATION_THRESHOLD

            if exceeds_threshold:
                # Population > 2000: Apply IA ratio to excess population
                excess_population = population - MIN_RURAL_POPULATION  # population - 2000

                # Distribute excess by IA ratio
                hab_urban = round(excess_population * urban_percentage)
                hab_rural_extra = round(excess_population * (1 - urban_percentage))
                hab_rural = MIN_RURAL_POPULATION + hab_rural_extra  # 2000 + extra rural

                # Calculate equipment
                # Base: fixed 2000 / 51
                base_equipos = MIN_RURAL_POPULATION / EQUIPMENT_DIVISOR_RURAL
                # Urban portion: / 301
                equipos_urban = hab_urban / EQUIPMENT_DIVISOR_URBAN
                # Rural extra portion: / 51
                equipos_rural_extra = hab_rural_extra / EQUIPMENT_DIVISOR_RURAL

                # Total = base + urban + rural_extra
                equipos_rural = round(base_equipos + equipos_rural_extra, 2)
                equipos_urban = round(equipos_urban, 2)
            else:
                # Population <= 2000: 100% rural (fixed)
                hab_urban = None  # Empty cell
                hab_rural = population

                # Calculate equipment
                equipos_urban = 0
                equipos_rural = round(population / EQUIPMENT_DIVISOR_RURAL, 2)

            # Total equipment is the sum
            total_equipos = round((equipos_urban or 0) + equipos_rural, 2)

            processed.append({
                'name': name,
                'total_hab': population,
                'hab_urban': hab_urban,
                'hab_rural': hab_rural,
                'equipos_urban': equipos_urban,
                'equipos_rural': equipos_rural,
                'total_equipos': total_equipos
            })

        return processed

    def create_full_excel(self, municipalities, output_path=None):
        """
        Create full Excel file with 7 columns

        Args:
            municipalities: List of municipality dictionaries
            output_path: Output file path (optional, uses config default)

        Returns:
            Path to created file
        """
        self._ensure_data_dir()
        output_path = output_path or self.output_full

        # Calculate equipment data
        processed_data = self._calculate_equipment_data(municipalities)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Municipios de {self.country_name}"

        # Write header
        headers = [
            self.column_names['name'],
            self.column_names['total_hab'],
            self.column_names['hab_urban'],
            self.column_names['hab_rural'],
            self.column_names['equipos_urban'],
            self.column_names['equipos_rural'],
            self.column_names['total_equipos']
        ]

        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.border = self.border

        # Write data
        for r_idx, m in enumerate(processed_data, 2):
            ws.cell(row=r_idx, column=1, value=m['name']).border = self.border
            ws.cell(row=r_idx, column=2, value=m['total_hab']).border = self.border
            ws.cell(row=r_idx, column=3, value=m['hab_urban']).border = self.border
            ws.cell(row=r_idx, column=4, value=m['hab_rural']).border = self.border
            ws.cell(row=r_idx, column=5, value=m['equipos_urban']).border = self.border
            ws.cell(row=r_idx, column=6, value=m['equipos_rural']).border = self.border
            ws.cell(row=r_idx, column=7, value=m['total_equipos']).border = self.border

        # Apply styles
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save
        wb.save(output_path)
        print(f"Full Excel saved: {output_path}")

        return output_path

    def create_simple_excel(self, municipalities, output_path=None):
        """
        Create simplified Excel file with only 2 columns

        Args:
            municipalities: List of municipality dictionaries
            output_path: Output file path (optional, uses config default)

        Returns:
            Path to created file
        """
        self._ensure_data_dir()
        output_path = output_path or self.output_simple

        # Calculate equipment data
        processed_data = self._calculate_equipment_data(municipalities)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Equipos"

        # Write header
        ws.cell(row=1, column=1, value=self.simple_columns['name']).border = self.border
        ws.cell(row=1, column=2, value=self.simple_columns['total_equipos']).border = self.border

        # Write data
        for r_idx, m in enumerate(processed_data, 2):
            ws.cell(row=r_idx, column=1, value=m['name']).border = self.border
            ws.cell(row=r_idx, column=2, value=m['total_equipos']).border = self.border

        # Apply styles
        self._apply_header_style(ws)
        self._auto_adjust_columns(ws)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save
        wb.save(output_path)
        print(f"Simple Excel saved: {output_path}")

        return output_path

    def create_both_excels(self, municipalities):
        """
        Create both full and simple Excel files

        Args:
            municipalities: List of municipality dictionaries

        Returns:
            Tuple of (full_path, simple_path)
        """
        full_path = self.create_full_excel(municipalities)
        simple_path = self.create_simple_excel(municipalities)

        return full_path, simple_path


def main():
    """Test the Excel generator"""
    generator = ExcelGenerator()

    # Test data with IA classification ratios (urban_percentage)
    # urban_percentage: ratio from IA classification (e.g., 0.7 = 70% urbano, 30% rural)
    test_municipalities = [
        {'name': 'Tudela', 'population': 37008, 'urban_percentage': 0.70},      # > 2000, 70% urbano
        {'name': 'Tafalla', 'population': 10582, 'urban_percentage': 0.80},     # > 2000, 80% urbano
        {'name': 'Sartaguda', 'population': 1287},    # <= 2000, 100% rural
        {'name': 'Sesma', 'population': 1149},        # <= 2000, 100% rural
        {'name': 'Sorlada', 'population': 51},        # <= 2000, 100% rural
        {'name': 'Ulzama', 'population': 1669},       # <= 2000, 100% rural
    ]

    generator.create_both_excels(test_municipalities)
    print("\nTest Excel files created successfully!")
    print(f"Urban divisor: {EQUIPMENT_DIVISOR_URBAN}")
    print(f"Rural divisor: {EQUIPMENT_DIVISOR_RURAL}")
    print(f"Default urban %: {DEFAULT_URBAN_PERCENTAGE * 100}%")
    print("\nNew Logic:")
    print("- Population <= 2000: 100% rural, equipos = population / 51")
    print("- Population > 2000:")
    print("  - Base 2000 (fijo) -> base_equipos = 2000 / 51")
    print("  - Excess distributed by IA ratio (urbano/rural)")
    print("  - Urbano: excess * urban% / 301")
    print("  - Rural extra: excess * rural% / 51")


if __name__ == "__main__":
    main()
